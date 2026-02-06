import pandas as pd
import sqlite3
import random
import os
import re
import csv
import io
import math
import threading
import time
import requests
import json
import string
import hmac
import hashlib
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, send_from_directory, jsonify, redirect, url_for, current_app, Response, send_file
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_mail import Message # NEW: For Email
from .models import User, get_db, SenderAddress
from .extensions import limiter, mail # NEW: Added mail
from .services import parser 
from .services.label_engine import LabelEngine
from .services.amazon_confirmer import run_confirmation, parse_cookies_and_csrf, validate_session

main_bp = Blueprint('main', __name__)

# --- CONFIGURATION ---
STRICT_HEADERS = [
    'No', 'FromName', 'PhoneFrom', 'Street1From', 'CompanyFrom', 'Street2From', 
    'CityFrom', 'StateFrom', 'PostalCodeFrom', 'ToName', 'PhoneTo', 'Street1To', 
    'Company2', 'Street2To', 'CityTo', 'StateTo', 'ZipTo', 'Weight', 'Length', 
    'Width', 'Height', 'Description', 'Ref01', 'Ref02', 'Contains Hazard', 'Shipment Date'
]

# SECURITY: LOAD FROM ENV
OXAPAY_KEY = os.getenv('OXAPAY_KEY')
OXAPAY_WEBHOOK_SECRET = os.getenv('OXAPAY_WEBHOOK_SECRET')
ACTIVE_CONFIRMATIONS = set() 

# --- SECURITY LOCKS ---
processing_lock = threading.Lock()
address_lock = threading.Lock()
deposit_lock = threading.Lock() # [SECURITY FIX] Lock for deposit transactions

# --- HELPER: LOGGING ---
def log_debug(message):
    print(f"[{datetime.now()}] [ROUTES] {message}")

# --- HELPER: DB CONNECTION ---
def get_db_conn():
    """Opens a DB connection with a high timeout to prevent locking errors."""
    conn = sqlite3.connect(current_app.config['DB_PATH'], timeout=60) # High Load Optimization
    conn.execute("PRAGMA journal_mode=WAL") # Write-Ahead Logging
    return conn

# --- HELPER: SANITIZATION (PREVENT ZPL INJECTION) ---
def sanitize_input(text, max_len=120):
    if text is None: return ""
    # Remove ZPL command markers, control chars, and cap field length.
    cleaned = str(text).replace('^', '').replace('~', '')
    cleaned = ''.join(ch for ch in cleaned if ch.isprintable())
    cleaned = re.sub(r'[\r\n\t]+', ' ', cleaned).strip()
    return cleaned[:max_len]

# --- HELPER: SEND OTP EMAIL ---
def send_otp_email(user):
    otp = ''.join(random.choices(string.digits, k=6)) # Generate 6 digit code
    
    conn = get_db_conn()
    c = conn.cursor()
    # Save OTP to DB with TIMESTAMP
    c.execute("UPDATE users SET otp_code = ?, otp_created_at = ? WHERE id = ?", 
              (otp, datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"), user.id))
    conn.commit()
    conn.close()

    try:
        msg = Message(f"Verification Code: {otp}", recipients=[user.email])
        msg.body = f"Your verification code is: {otp}\n\nThis code expires in 10 minutes."
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Email Error: {e}")
        return False

# --- HELPER: DATAFRAME ---
def normalize_dataframe(df):
    df.columns = [str(c).strip() for c in df.columns]
    
    # --- FIX: Check for required columns BEFORE accessing them ---
    required_cols = ['ToName', 'Street1To']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        return None, f"Invalid CSV Format. Missing columns: {', '.join(missing_cols)}"

    if 'No' in df.columns:
        if df['No'].isnull().any() or (df['No'] == '').any():
            return None, "Row Error: Missing Order Number in 'No' column."

    # Now safe to access columns
    incomplete_rows = df[df['ToName'].isnull() | (df['ToName'].astype(str).str.strip() == '') | 
                          df['Street1To'].isnull() | (df['Street1To'].astype(str).str.strip() == '')]
    if not incomplete_rows.empty:
        first_error_idx = incomplete_rows.index[0] + 2
        return None, f"Row {first_error_idx} Error: Missing required Recipient Information."

    if 'StateTo' in df.columns:
        df['StateTo'] = df['StateTo'].astype(str).str.strip().str.upper()
        bad_states = df[df['StateTo'].str.len() != 2]
        if not bad_states.empty:
            bad_row = bad_states.index[0] + 2
            bad_val = bad_states.iloc[0]['StateTo']
            return None, f"Row {bad_row} Error: State '{bad_val}' must be 2-letter code (e.g. 'CA')."

    if 'StateFrom' in df.columns:
        df['StateFrom'] = df['StateFrom'].astype(str).str.strip().str.upper()
        bad_from = df[df['StateFrom'].str.len() != 2]
        if not bad_from.empty:
             bad_row = bad_from.index[0] + 2
             bad_val = bad_from.iloc[0]['StateFrom']
             return None, f"Row {bad_row} Error: Sender State '{bad_val}' must be 2 letters."

    if 'ZipTo' in df.columns:
        df['ZipTo'] = df['ZipTo'].astype(str).str.split('.').str[0].str.strip()
        short_zips = df[df['ZipTo'].str.len() < 5]
        if not short_zips.empty:
            bad_row = short_zips.index[0] + 2
            bad_val = short_zips.iloc[0]['ZipTo']
            return None, f"Row {bad_row} Error: Zip Code '{bad_val}' is too short."

    return df, None

def get_system_config():
    conn = get_db_conn(); c = conn.cursor()
    try:
        c.execute("SELECT key, value FROM system_config")
        rows = dict(c.fetchall())
    except: rows = {}
    conn.close()
    return rows

def get_price(user_id, label_type, version, default_price):
    try:
        conn = get_db_conn(); c = conn.cursor()
        c.execute("SELECT price FROM user_pricing WHERE user_id = ? AND label_type = ? AND version = ?", (user_id, label_type, version))
        row = c.fetchone(); conn.close()
        if row: return float(row[0])
    except: pass
    return default_price

def to_est(date_str):
    try:
        if not date_str: return ""
        dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        est = dt - timedelta(hours=5)
        return est.strftime("%Y-%m-%d %H:%M:%S")
    except: return date_str

# --- HELPER: CHECK ENABLED VERSIONS ---
def get_enabled_versions():
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT key, value FROM system_config WHERE key LIKE 'ver_en_%'")
    rows = dict(c.fetchall())
    conn.close()
    
    return {
        '95055': rows.get('ver_en_95055', '1') == '1',
        '94888': rows.get('ver_en_94888', '1') == '1',
        '94019': rows.get('ver_en_94019', '1') == '1',
        '95888': rows.get('ver_en_95888', '1') == '1',
        '91149': rows.get('ver_en_91149', '1') == '1',
        '93055': rows.get('ver_en_93055', '1') == '1'
    }

def is_version_enabled(version):
    status = get_enabled_versions()
    return status.get(version, True)

# --- AUTH ROUTES ---
@main_bp.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.is_admin: return redirect(url_for('admin.dashboard'))
        # UPDATE: Redirect to Single Label instead of Purchase
        return redirect(url_for('main.single'))
    return redirect(url_for('main.login'))

@main_bp.route('/login', methods=['GET', 'POST'])
@limiter.limit("50 per minute")
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        remember = True if request.form.get('remember') else False
        
        user_data = User.get_by_username(username)
        
        if user_data and check_password_hash(user_data[3], password):
            # [SECURITY FIX] Check if banned
            if user_data[7]: # is_banned index
                return render_template('login.html', error="ACCOUNT SUSPENDED")

            # NEW: Check if verified
            conn = get_db_conn()
            c = conn.cursor()
            c.execute("SELECT is_verified, email FROM users WHERE id = ?", (user_data[0],))
            res = c.fetchone()
            conn.close()
            
            # If is_verified is 0 or NULL
            if not res or not res[0]:
                user_obj = User.get(user_data[0])
                send_otp_email(user_obj) # Resend OTP
                return render_template('verify.html', email=res[1], error="Unverified Device. Code sent.")

            user = User.get(user_data[0]) 
            login_user(user, remember=remember)
            
            try:
                conn = get_db_conn(); c = conn.cursor()
                ip = request.headers.get('X-Forwarded-For', request.remote_addr)
                ua = request.headers.get('User-Agent', '')[:200]
                c.execute("INSERT INTO login_history (user_id, ip_address, user_agent, created_at) VALUES (?, ?, ?, ?)",
                          (user.id, ip, ua, datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")))
                conn.commit(); conn.close()
            except: pass
            
            if user.is_admin: return redirect(url_for('admin.dashboard'))
            # UPDATE: Redirect to Single Label
            return redirect(url_for('main.single'))
        return render_template('login.html', error="INVALID CREDENTIALS")
    return render_template('login.html')

@main_bp.route('/register', methods=['GET', 'POST'])
@limiter.limit("5 per minute") # Global limit to stop spam
def register():
    if request.method == 'POST':
        username = request.form['username']; email = request.form['email']; password = request.form['password']
        hashed = generate_password_hash(password)
        
        # 1. Create User (Unverified by default)
        user = User.create(username, email, hashed)
        if user:
            # 2. Send 2FA Code
            if send_otp_email(user):
                return render_template('verify.html', email=email, message="Code sent! Check your inbox.")
            else:
                return render_template('login.html', mode="register", error="EMAIL FAILED")
        else: 
            return render_template('login.html', mode="register", error="USERNAME/EMAIL TAKEN")
    return render_template('login.html', mode="register")

# --- NEW: RESEND CODE ROUTE (WITH 90s COOLDOWN) ---
@main_bp.route('/resend_code', methods=['POST'])
@limiter.limit("5 per minute")
def resend_code():
    email = request.form.get('email')
    if not email: return jsonify({"status": "error", "message": "Email missing"})
    
    conn = get_db_conn()
    c = conn.cursor()
    c.execute("SELECT id, username, email, otp_created_at FROM users WHERE email = ?", (email,))
    row = c.fetchone()
    
    if not row: 
        conn.close()
        return jsonify({"status": "error", "message": "User not found"})
    
    user_id, username, user_email, db_time = row

    # --- CHECK 90s COOLDOWN ---
    if db_time:
        try:
            last_sent = datetime.strptime(db_time, "%Y-%m-%d %H:%M:%S")
            if (datetime.utcnow() - last_sent).total_seconds() < 90:
                conn.close()
                return jsonify({"status": "error", "message": "Please wait 90s before resending"})
        except: pass
    
    conn.close()

    # Create temp user object for email function
    class TempUser:
        def __init__(self, uid, em):
            self.id = uid
            self.email = em
    
    if send_otp_email(TempUser(user_id, user_email)):
        return jsonify({"status": "success", "message": "Code Resent"})
    else:
        return jsonify({"status": "error", "message": "Email failed"})

@main_bp.route('/verify', methods=['POST'])
@limiter.limit("10 per minute")
def verify_account():
    email = request.form['email']
    code = request.form['code']
    
    # We must fetch manually because User.get_by_username doesn't expose fields easily via ORM
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT id, otp_code, otp_created_at, is_verified FROM users WHERE email = ?", (email,))
    row = c.fetchone()
    conn.close()
    
    if not row: return render_template('verify.html', error="User not found")
    
    user_id, db_code, db_time, is_ver = row
    
    if is_ver: return redirect(url_for('main.login'))
    
    # [SECURITY FIX] Check Code and Expiry (10 mins)
    valid_code = str(db_code) == str(code)
    not_expired = True
    if db_time:
        try:
            created_at = datetime.strptime(db_time, "%Y-%m-%d %H:%M:%S")
            if (datetime.utcnow() - created_at).total_seconds() > 600: # 10 mins
                not_expired = False
        except: pass

    if valid_code and not_expired:
        conn = get_db_conn(); c = conn.cursor()
        c.execute("UPDATE users SET is_verified = 1, otp_code = NULL WHERE id = ?", (user_id,))
        conn.commit(); conn.close()
        
        user = User.get(user_id)
        login_user(user)
        
        # --- FIXED: CHECK FOR ADMIN STATUS AND REDIRECT ---
        if user.is_admin: return redirect(url_for('admin.dashboard'))
        
        # UPDATE: Redirect to Single Label
        return redirect(url_for('main.single'))
    elif not not_expired:
        return render_template('verify.html', email=email, error="Code Expired (10m)")
    else:
        return render_template('verify.html', email=email, error="Invalid Code")

@main_bp.route('/logout')
@login_required
def logout(): logout_user(); return redirect(url_for('main.login'))

# --- DASHBOARD ROUTES ---
@main_bp.route('/dashboard')
@login_required
def dashboard_root(): 
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    # UPDATE: Redirect to Single Label
    return redirect(url_for('main.single'))

@main_bp.route('/purchase')
@login_required
def purchase(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    
    # [WALMART] Added addresses for dropdown
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT * FROM sender_addresses WHERE user_id = ?", (current_user.id,))
    rows = c.fetchall()
    conn.close()
    addrs = [{"id": r[0], "name": r[2], "street1": r[5]} for r in rows]
    return render_template('dashboard.html', user=current_user, active_tab='purchase', version_status=get_enabled_versions(), addresses=addrs)

# --- NEW: SINGLE PURCHASE ROUTE ---
@main_bp.route('/single')
@login_required
def single(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    
    # Fetch addresses just like /purchase does
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT * FROM sender_addresses WHERE user_id = ?", (current_user.id,))
    rows = c.fetchall()
    conn.close()
    addrs = [{"id": r[0], "name": r[2], "street1": r[5]} for r in rows]
    return render_template('dashboard.html', user=current_user, active_tab='single', version_status=get_enabled_versions(), addresses=addrs)

@main_bp.route('/history')
@login_required
def history(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    return render_template('dashboard.html', user=current_user, active_tab='history', version_status=get_enabled_versions())

@main_bp.route('/automation')
@login_required
def automation(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    
    sys_config = get_system_config()
    monthly_left = int(sys_config.get('slots_monthly_total', 50)) - int(sys_config.get('slots_monthly_used', 0))
    lifetime_left = int(sys_config.get('slots_lifetime_total', 10)) - int(sys_config.get('slots_lifetime_used', 0))
    p_month = sys_config.get('automation_price_monthly', '29.99')
    p_life = sys_config.get('automation_price_lifetime', '499.00')
    return render_template('dashboard.html', user=current_user, active_tab='automation', monthly_left=monthly_left, lifetime_left=lifetime_left, price_monthly=p_month, price_lifetime=p_life, system_status="OPERATIONAL", version_status=get_enabled_versions())

# --- NEW: INVENTORY ROUTE ---
@main_bp.route('/inventory')
@login_required
def inventory(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    return render_template('dashboard.html', user=current_user, active_tab='inventory', version_status=get_enabled_versions())

@main_bp.route('/stats')
@login_required
def stats(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    return render_template('dashboard.html', user=current_user, active_tab='stats', version_status=get_enabled_versions())

@main_bp.route('/deposit')
@login_required
def deposit():
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    return render_template('dashboard.html', user=current_user, active_tab='deposit', version_status=get_enabled_versions())

@main_bp.route('/settings')
@login_required
def settings(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    return render_template('dashboard.html', user=current_user, active_tab='settings', version_status=get_enabled_versions())

@main_bp.route('/addresses')
@login_required
def addresses(): 
    # --- ADMIN REDIRECT ---
    if current_user.is_admin: return redirect(url_for('admin.dashboard'))
    return render_template('dashboard.html', user=current_user, active_tab='addresses', version_status=get_enabled_versions())

# --- API ENDPOINTS ---
@main_bp.route('/api/user')
@login_required
def api_user():
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT version, price FROM user_pricing WHERE user_id = ? AND label_type = 'priority'", (current_user.id,))
    prices = {row[0]: float(row[1]) for row in c.fetchall()}
    conn.close()
    base = current_user.price_per_label
    return jsonify({
        "username": current_user.username, 
        "balance": current_user.balance, 
        "price_per_label": base,
        "prices": { 
            "95055": prices.get('95055', base), 
            "94888": prices.get('94888', base), 
            "94019": prices.get('94019', base),
            "95888": prices.get('95888', base), 
            "91149": prices.get('91149', base), 
            "93055": prices.get('93055', base)
        }
    })

@main_bp.route('/api/notifications/poll')
@login_required
def poll_notifications():
    try:
        conn = get_db_conn(); c = conn.cursor()
        c.execute("SELECT id, message, type FROM user_notifications WHERE user_id = ?", (current_user.id,))
        rows = c.fetchall()
        notifs = [{"id": r[0], "msg": r[1], "type": r[2]} for r in rows]
        if notifs:
            ids = [str(n['id']) for n in notifs]
            c.execute(f"DELETE FROM user_notifications WHERE id IN ({','.join(ids)})")
            conn.commit()
        conn.close()
        return jsonify(notifs)
    except: return jsonify([])

@main_bp.route('/api/settings/defaults', methods=['POST'])
@login_required
def save_defaults():
    data = request.json; current_user.update_defaults(data.get('label_type','priority'), data.get('version','95055'), data.get('template','pitney_v2'))
    return jsonify({"status": "success", "message": "DEFAULTS SAVED"})

@main_bp.route('/api/batches')
@login_required
def api_batches():
    page = int(request.args.get('page', 1)); limit = 10; offset = (page-1)*limit
    search = request.args.get('search', '').strip(); sort_by = request.args.get('sort', 'recent')
    view = request.args.get('view', '') 
    
    conn = get_db_conn(); c = conn.cursor()
    
    if view == 'history':
        query = "SELECT * FROM batches WHERE user_id = ?"
    else:
        query = "SELECT * FROM batches WHERE user_id = ? AND filename NOT LIKE 'WALMART_%'"
        
    params = [current_user.id]
    
    if search: query += " AND (batch_id LIKE ? OR filename LIKE ?)"; params.extend([f"%{search}%", f"%{search}%"])
    query += " ORDER BY created_at ASC" if sort_by == 'oldest' else " ORDER BY count DESC" if sort_by == 'high' else " ORDER BY created_at DESC"
    
    c.execute(query.replace("SELECT *", "SELECT COUNT(*)"), params); total_items = c.fetchone()[0]
    query += " LIMIT ? OFFSET ?"; params.extend([limit, offset])
    c.execute(query, params); rows = c.fetchall(); conn.close()
    
    data = []
    for r in rows:
        b_id=r[0]; est_date=to_est(r[9]); is_exp=False
        try: 
            if (datetime.utcnow() - datetime.strptime(r[9], "%Y-%m-%d %H:%M:%S")).days >= 7: is_exp = True
        except: pass
        
        filename_val = r[2]
        is_walmart_batch = filename_val.startswith("WALMART_")
        clean_display_name = filename_val.split('_',1)[1] if '_' in filename_val else filename_val
        
        data.append({
            "batch_id": b_id, 
            "batch_name": clean_display_name, 
            "count": r[3], 
            "success_count": r[4], 
            "status": r[5], 
            "date": est_date, 
            "is_expired": is_exp,
            "is_walmart": is_walmart_batch
        })
    return jsonify({"data": data, "pagination": {"current_page": page, "total_pages": math.ceil(total_items/limit)}})

@main_bp.route('/api/stats')
@login_required
def api_stats():
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT SUM(success_count) FROM batches WHERE user_id = ?", (current_user.id,)); total = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM batches WHERE user_id = ?", (current_user.id,)); batches = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM batches WHERE user_id = ? AND status = 'READY'", (current_user.id,)); ready = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM batches WHERE user_id = ? AND status = 'PARTIAL'", (current_user.id,)); partial = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM batches WHERE user_id = ? AND status IN ('FAILED', 'AUTH_ERROR')", (current_user.id,)); failed = c.fetchone()[0] or 0
    conn.close()
    return jsonify({"total_labels": total, "total_batches": batches, "ready_batches": ready, "partial_batches": partial, "failed_batches": failed})

@main_bp.route('/process', methods=['POST'])
@login_required
@limiter.limit("30 per minute") 
def process():
    if 'file' not in request.files: return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    if file.filename == '': return jsonify({"error": "No selected file"}), 400
    
    upload_mode = request.form.get('upload_mode', 'bulk') 
    req_version = request.form.get('tracking_version')
    if not is_version_enabled(req_version):
        return jsonify({"error": f"SERVICE UNAVAILABLE: Version {req_version} is currently disabled."}), 400

    # --- [SECURITY FIX] VALIDATE TEMPLATE INPUT ---
    req_template = request.form.get('template_choice')
    allowed_templates = ['pitney_v2', 'stamps_v2', 'easypost_v2']
    if req_template not in allowed_templates:
         return jsonify({"error": "Invalid Template Selection"}), 400

    log_debug(f"Processing Upload: {file.filename} Mode: {upload_mode}")
    price = get_price(current_user.id, request.form.get('label_type'), req_version, current_user.price_per_label)

    # --- WALMART LOGIC ---
    if upload_mode == 'walmart':
        sender_id = request.form.get('sender_id')
        if not sender_id: return jsonify({"error": "Sender Profile Required for Walmart Mode"}), 400
        sender = SenderAddress.get(sender_id)
        if not sender or sender.user_id != current_user.id: return jsonify({"error": "Invalid Sender Profile"}), 400
        
        data, count, _ = parser.parse_walmart_xlsx(file, sender)
        if count == 0: return jsonify({"error": "No valid rows found in XLSX. Check format."}), 400
        
        df = pd.DataFrame(data)
        mapping = {
            'to_name': 'ToName', 'to_phone': 'PhoneTo', 'to_street1': 'Street1To', 'to_street2': 'Street2To', 
            'to_city': 'CityTo', 'to_state': 'StateTo', 'to_zip': 'ZipTo',
            'from_name': 'FromName', 'from_phone': 'PhoneFrom', 'from_street1': 'Street1From', 
            'from_company': 'CompanyFrom', 'from_street2': 'Street2From', 'from_city': 'CityFrom', 
            'from_state': 'StateFrom', 'from_zip': 'PostalCodeFrom',
            'weight': 'Weight', 'reference': 'Description'
        }
        df.rename(columns=mapping, inplace=True)
        
        df['No'] = range(1, len(df) + 1)
        df['Length'] = 10; df['Width'] = 6; df['Height'] = 4
        df['Contains Hazard'] = 'False'
        df['Shipment Date'] = datetime.now().strftime("%m/%d/%Y")
        
        for col in STRICT_HEADERS:
            if col not in df.columns: df[col] = ''
        df = df[STRICT_HEADERS]
        
        file_prefix = "WALMART_"
        file.stream.seek(0)
        
    else:
        try:
            file.stream.seek(0)
            df = pd.read_csv(file, encoding='utf-8-sig', on_bad_lines='skip', dtype=str)
        except Exception as e:
            log_debug(f"CSV Read Fail: {e}")
            return jsonify({"error": "Could not read CSV file. Please ensure it is a valid CSV format."}), 400

        df, error_msg = normalize_dataframe(df)
        if error_msg: return jsonify({"error": error_msg}), 400
        file_prefix = ""

    cost = len(df) * price
    
    # --- SECURITY LOCK START (Prevent Balance Race Conditions) ---
    with processing_lock:
        if not current_user.update_balance(-cost): 
            return jsonify({"error": "INSUFFICIENT FUNDS"}), 402
        
        try:
            batch_id = None
            for _ in range(10): 
                temp_id = str(random.randint(100000, 999999))
                conn = get_db_conn(); c = conn.cursor()
                c.execute("SELECT 1 FROM batches WHERE batch_id = ?", (temp_id,))
                exists = c.fetchone(); conn.close()
                if not exists: batch_id = temp_id; break
            
            if not batch_id:
                current_user.update_balance(cost) 
                return jsonify({"error": "System Busy: Could not allocate Batch ID. Please try again."}), 500

            clean_name = secure_filename(file.filename)
            if not clean_name: clean_name = "upload.csv"
            if clean_name.lower().endswith('.xlsx'): clean_name = clean_name[:-5] + ".csv"
            
            final_filename = f"{file_prefix}{batch_id}_{clean_name}"
            save_path = os.path.join(current_app.config['DATA_FOLDER'], 'uploads', final_filename)
            df.to_csv(save_path, index=False)
            
            if upload_mode == 'walmart':
                orig_name = f"{file_prefix}{batch_id}_ORIG.xlsx"
                orig_path = os.path.join(current_app.config['DATA_FOLDER'], 'uploads', orig_name)
                file.stream.seek(0)
                file.save(orig_path)
            
            conn = get_db_conn(); c = conn.cursor()
            c.execute("INSERT INTO batches (batch_id, user_id, filename, count, success_count, status, template, version, label_type, created_at, price) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                      (batch_id, current_user.id, final_filename, len(df), 0, 'QUEUED', req_template, req_version, request.form.get('label_type'), datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"), price))
            conn.commit(); conn.close()
            log_debug(f"Batch {batch_id} Queued Successfully")
            return jsonify({"status": "success", "batch_id": batch_id})
        except Exception as e:
            log_debug(f"[CRITICAL ERROR] /process endpoint: {str(e)}")
            current_user.update_balance(cost) # Refund on crash
            return jsonify({"error": "Internal System Error. Please try again later."}), 500
    # --- SECURITY LOCK END ---

@main_bp.route('/verify-csv', methods=['POST'])
@login_required
@limiter.limit("20 per minute")
def verify_csv():
    if 'file' not in request.files: return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    upload_mode = request.form.get('upload_mode', 'bulk')
    req_version = request.form.get('tracking_version', '95055')
    if not is_version_enabled(req_version): return jsonify({"error": f"SERVICE UNAVAILABLE: Version {req_version} is currently disabled."}), 400
    
    count = 0
    if upload_mode == 'walmart':
        sender_id = request.form.get('sender_id')
        if not sender_id: return jsonify({"error": "Sender Profile Required"}), 400
        sender = SenderAddress.get(sender_id)
        if not sender or sender.user_id != current_user.id: return jsonify({"error": "Invalid Sender"}), 400
        _, count, _ = parser.parse_walmart_xlsx(file, sender)
        if count == 0: return jsonify({"error": "No valid rows found in XLSX"}), 400
    else:
        try: 
            file.stream.seek(0)
            df = pd.read_csv(file, encoding='utf-8-sig', on_bad_lines='skip', dtype=str)
        except Exception as e: return jsonify({"error": "Invalid CSV Format"}), 400
        
        # --- FIXED CALL HERE ---
        df, error_msg = normalize_dataframe(df)
        if error_msg: return jsonify({"error": error_msg}), 400
        count = len(df)

    label_type = request.form.get('label_type', 'priority')
    price = get_price(current_user.id, label_type, req_version, current_user.price_per_label)
    return jsonify({"count": count, "cost": count * price})

# --- DOWNLOAD ROUTES ---
@main_bp.route('/api/download/xlsx/<batch_id>')
@login_required
def download_xlsx(batch_id):
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT filename, status FROM batches WHERE batch_id = ? AND user_id = ?", (batch_id, current_user.id)); batch_row = c.fetchone()
    if not batch_row: conn.close(); return jsonify({"error": "UNAUTHORIZED"}), 403
    if not batch_row[0].startswith("WALMART_"): conn.close(); return jsonify({"error": "NOT A WALMART BATCH"}), 400
    
    c.execute("SELECT ref02, tracking FROM history WHERE batch_id = ? AND status = 'SUCCESS'", (batch_id,))
    tracking_map = {str(row[0]).strip(): row[1] for row in c.fetchall()}
    conn.close()

    orig_name = f"WALMART_{batch_id}_ORIG.xlsx"
    path = os.path.join(current_app.config['DATA_FOLDER'], 'uploads', orig_name)
    if not os.path.exists(path): return jsonify({"error": "ORIGINAL FILE NOT FOUND"}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                order_num_cell = row[1].value
                qty_cell = row[24].value

                order_id = str(order_num_cell).strip()
                if order_id.endswith('.0'): order_id = order_id[:-2]

                tracking = tracking_map.get(order_id)

                if tracking:
                    sheet.cell(row=row_idx, column=40).value = "Ship"
                    sheet.cell(row=row_idx, column=41).value = qty_cell
                    sheet.cell(row=row_idx, column=42).value = "USPS"
                    sheet.cell(row=row_idx, column=43).value = tracking
                    sheet.cell(row=row_idx, column=44).value = "www.usps.com"
            except Exception:
                continue

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"Walmart_Fulfilled_{batch_id}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"XLSX Gen Error: {e}")
        return jsonify({"error": "FAILED TO GENERATE FILE"}), 500

# --- OXAPAY: WEBHOOK SIGNATURE CHECK ---
def is_valid_oxapay_webhook(req):
    if not OXAPAY_WEBHOOK_SECRET:
        print("[PAYMENT] WARNING: OXAPAY_WEBHOOK_SECRET not set; webhook signature not enforced")
        return True

    provided = (
        req.headers.get('X-OxaPay-Signature')
        or req.headers.get('X-OXAPAY-SIGNATURE')
        or req.headers.get('X-Signature')
        or ''
    ).strip()
    if not provided:
        return False

    raw_body = req.get_data() or b''
    expected = hmac.new(OXAPAY_WEBHOOK_SECRET.encode('utf-8'), raw_body, hashlib.sha256).hexdigest()
    return hmac.compare_digest(provided.lower(), expected.lower())

# --- OXAPAY: VERIFY HELPER ---
def verify_oxapay_payment(track_id):
    try:
        if not OXAPAY_KEY: 
            print("[PAYMENT] Missing OXAPAY_KEY")
            return False, 0.0, "Missing API Key"
            
        url = "https://api.oxapay.com/merchants/inquiry"
        payload = {"merchant": OXAPAY_KEY, "trackId": track_id}
        r = requests.post(url, json=payload, timeout=10)
        data = r.json()
        
        if data.get('result') == 100:
            status = data.get('status', '').lower()
            if status in ['paid', 'complete']: 
                usd_val = float(data.get('amount', 0))
                return True, usd_val, "Paid"
            else:
                return False, 0.0, f"Status: {status.upper()}"
        else:
            return False, 0.0, f"Gateway Error: {data.get('message', 'Unknown')}"
            
    except Exception as e: 
        print(f"[PAYMENT ERROR] {e}")
        return False, 0.0, "Connection Error"

@main_bp.route('/api/deposit/history', methods=['GET'])
@login_required
def get_deposit_history():
    conn = get_db_conn(); c = conn.cursor()
    try: c.execute("UPDATE deposit_history SET status='FAILED' WHERE status='PROCESSING' AND created_at < ?", ((datetime.utcnow()-timedelta(minutes=60)).strftime("%Y-%m-%d %H:%M:%S"),)); conn.commit()
    except: pass
    c.execute("SELECT amount, currency, txn_id, status, created_at FROM deposit_history WHERE user_id = ? ORDER BY id DESC LIMIT 10", (current_user.id,))
    data = [{"amount":r[0],"currency":r[1],"txn_id":r[2],"status":r[3],"date":to_est(r[4])} for r in c.fetchall()]
    conn.close(); return jsonify(data)

@main_bp.route('/api/deposit/check/<txn_id>', methods=['POST'])
@login_required
def manual_check_deposit(txn_id):
    with deposit_lock:
        conn = get_db_conn(); c = conn.cursor()
        c.execute("SELECT id, status FROM deposit_history WHERE txn_id = ? AND user_id = ?", (txn_id, current_user.id))
        row = c.fetchone()

        if not row:
            conn.close()
            return jsonify({"error": "Transaction not found"}), 404

        if row[1] == 'PAID':
            conn.close()
            return jsonify({"status": "success", "message": "Already Paid"})

        is_valid, paid_amount, status_msg = verify_oxapay_payment(txn_id)

        if is_valid and paid_amount > 0:
            c.execute("UPDATE deposit_history SET status='PAID', amount=? WHERE id=? AND status != 'PAID'", (paid_amount, row[0]))
            if c.rowcount > 0:
                c.execute("UPDATE users SET balance = balance + ? WHERE id = ?", (paid_amount, current_user.id))
                conn.commit(); conn.close()
                return jsonify({"status": "success", "message": f"Payment Confirmed! +${paid_amount} Added."})
            conn.close()
            return jsonify({"status": "success", "message": "Already Paid"})

        if status_msg and "Status:" in status_msg:
            clean_status = status_msg.split(': ')[1].strip()
            if clean_status in ['EXPIRED', 'FAILED']:
                c.execute("UPDATE deposit_history SET status=? WHERE id=?", (clean_status, row[0]))
                conn.commit()

        conn.close()
        return jsonify({"error": f"Gateway Report: {status_msg}"}), 400

@main_bp.route('/api/deposit/create', methods=['POST'])
@login_required
def create_deposit():
    if not OXAPAY_KEY: return jsonify({"error": "Payment Gateway Configuration Error"}), 503
    data = request.json; usd_amount = float(data.get('amount'))
    if usd_amount < 10: return jsonify({"error": "Minimum deposit is $10"}), 400
    url = "https://api.oxapay.com/merchants/request"
    custom_order_id = f"USER_{current_user.id}_{int(time.time())}_{usd_amount}"
    # UPDATE: Redirect to Single Label on return
    payload = {"merchant": OXAPAY_KEY, "amount": usd_amount, "currency": "USDT", "lifeTime": 60, "feePaidByPayer": 0, "underPaidCover": 2.0, "callbackUrl": url_for('main.deposit_webhook', _external=True), "returnUrl": url_for('main.single', _external=True), "description": f"Deposit ${usd_amount}", "orderId": custom_order_id}
    try:
        r = requests.post(url, json=payload); result = r.json()
        if result.get('result') == 100:
            track_id = result.get('trackId')
            conn = get_db_conn(); c = conn.cursor()
            c.execute("INSERT INTO deposit_history (user_id, amount, currency, txn_id, status, created_at) VALUES (?, ?, ?, ?, ?, ?)", (current_user.id, usd_amount, 'USDT', str(track_id), 'PROCESSING', datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit(); conn.close()
            return jsonify({"status": "success", "pay_link": result.get('payLink')})
        else: return jsonify({"error": "Gateway Unavailable"}), 400
    except: return jsonify({"error": "Connection Error"}), 500

@main_bp.route('/api/deposit/webhook', methods=['POST'])
def deposit_webhook():
    try:
        if not is_valid_oxapay_webhook(request):
            return jsonify({"status": "error", "message": "Invalid webhook signature"}), 403

        data = request.json or {}
        status = str(data.get('status', '')).lower().strip()
        order_id = str(data.get('orderId', '')).strip()
        track_id = str(data.get('trackId', '')).strip()

        if not track_id:
            return jsonify({"status": "error", "message": "Missing trackId"}), 400

        if status in ['paid', 'complete']:
            with deposit_lock:
                is_valid, verified_amount, _ = verify_oxapay_payment(track_id)
                if not (is_valid and verified_amount > 0):
                    return jsonify({"status": "ok"}), 200

                parts = order_id.split('_')
                if len(parts) < 2 or not parts[1].isdigit():
                    return jsonify({"status": "error", "message": "Invalid orderId"}), 400

                user_id = int(parts[1])
                conn = get_db_conn(); c = conn.cursor()
                c.execute("SELECT id, status FROM deposit_history WHERE txn_id = ?", (track_id,))
                existing = c.fetchone()

                if existing:
                    deposit_id, _current_status = existing
                    c.execute("UPDATE deposit_history SET status='PAID', currency=?, amount=? WHERE id=? AND status != 'PAID'", (data.get('currency', 'USDT'), verified_amount, deposit_id))
                    if c.rowcount > 0:
                        c.execute("UPDATE users SET balance = balance + ? WHERE id = ?", (verified_amount, user_id))
                        conn.commit()
                else:
                    c.execute("INSERT INTO deposit_history (user_id, amount, currency, txn_id, status, created_at) VALUES (?, ?, ?, ?, ?, ?)", (user_id, verified_amount, data.get('currency', 'USDT'), track_id, 'PAID', datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")))
                    c.execute("UPDATE users SET balance = balance + ? WHERE id = ?", (verified_amount, user_id))
                    conn.commit()

                conn.close()
                return jsonify({"status": "ok"}), 200

        elif status in ['expired', 'failed', 'rejected']:
            conn = get_db_conn(); c = conn.cursor()
            c.execute("UPDATE deposit_history SET status='FAILED' WHERE txn_id = ?", (track_id,)); conn.commit(); conn.close()

    except Exception as e:
        print(f"[WEBHOOK ERROR] {e}")
        return jsonify({"status": "error"}), 500

    return jsonify({"status": "ok"}), 200

# --- AUTOMATION PUBLIC ENDPOINTS ---
@main_bp.route('/api/automation/public_config')
@login_required
def public_automation_config():
    sys_config = get_system_config()
    monthly_left = int(sys_config.get('slots_monthly_total', 50)) - int(sys_config.get('slots_monthly_used', 0))
    lifetime_left = int(sys_config.get('slots_lifetime_total', 10)) - int(sys_config.get('slots_lifetime_used', 0))
    return jsonify({
        "monthly_left": monthly_left,
        "lifetime_left": lifetime_left,
        "monthly_price": sys_config.get('automation_price_monthly', '29.99'),
        "lifetime_price": sys_config.get('automation_price_lifetime', '499.00')
    })

@main_bp.route('/api/automation/save', methods=['POST'])
@login_required
def automation_save():
    if not current_user.is_subscribed: return jsonify({"error": "LICENSE REQUIRED"}), 403
    cookies = request.form.get('cookies', ''); csrf = request.form.get('csrf', '').strip(); inventory = request.form.get('inventory', '')
    current_user.update_settings(cookies, csrf, "", inventory, False)
    return jsonify({"status": "success", "message": "SETTINGS SAVED"})

@main_bp.route('/api/automation/format', methods=['POST'])
@login_required
@limiter.limit("100 per minute")
def automation_format():
    if not current_user.is_subscribed: return jsonify({"error": "LICENSE REQUIRED"}), 403
    file = request.files.get('file')
    if not file: return jsonify({"error": "NO FILE UPLOADED"}), 400
    address_id = request.form.get('address_id'); sender_address = None
    if address_id:
        conn = get_db_conn(); c = conn.cursor()
        c.execute("SELECT name, company, street1, street2, city, state, zip, phone FROM sender_addresses WHERE id = ? AND user_id = ?", (address_id, current_user.id))
        row = c.fetchone(); conn.close()
        if row: sender_address = {'name': row[0], 'company': row[1], 'street1': row[2], 'street2': row[3], 'city': row[4], 'state': row[5], 'zip': row[6], 'phone': row[7]}
    content = file.read()
    
    # --- FIX: USE CORRECT PARSER REFERENCE ---
    try:
        zip_bytes, error = parser.OrderParser.parse_to_zip(content, current_user.inventory_json, sender_address)
        if error: return jsonify({"error": error}), 400
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        return Response(zip_bytes, mimetype="application/zip", headers={"Content-disposition": f"attachment; filename=Parsed_Orders_{timestamp}.zip"})
    except Exception as e:
        return jsonify({"error": f"PARSER CRASH: {str(e)}"}), 500

@main_bp.route('/api/automation/confirm', methods=['POST'])
@login_required
@limiter.limit("120 per minute")
def automation_confirm():
    if not current_user.is_subscribed: return jsonify({"error": "UNAUTHORIZED: License Required"}), 403
    batch_id = request.json.get('batch_id')
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT status FROM batches WHERE batch_id = ? AND user_id = ?", (batch_id, current_user.id)); row = c.fetchone()
    if row and row[0] == 'REFUNDED': conn.close(); return jsonify({'error': 'ACCESS REVOKED: THIS BATCH WAS REFUNDED'}), 403
    conn.close()
    if batch_id in ACTIVE_CONFIRMATIONS: return jsonify({"error": "THIS BATCH IS ALREADY RUNNING"}), 429
    raw_cookies = current_user.auth_cookies; raw_csrf = current_user.auth_csrf
    if not raw_cookies or len(raw_cookies) < 10: return jsonify({"error": "MISSING COOKIES"}), 400
    final_cookies, final_csrf = parse_cookies_and_csrf(raw_cookies); real_csrf = final_csrf if final_csrf else raw_csrf
    is_valid, msg = validate_session(final_cookies, real_csrf)
    if not is_valid: return jsonify({"error": msg}), 400
    ACTIVE_CONFIRMATIONS.add(batch_id); db_path = current_app.config['DB_PATH']
    def task(app_context):
        with app_context:
            try:
                conn = sqlite3.connect(db_path, timeout=30); c = conn.cursor()
                c.execute("UPDATE batches SET status = 'CONFIRMING' WHERE batch_id = ?", (batch_id,)); conn.commit(); conn.close()
                run_confirmation(batch_id, raw_cookies, raw_csrf)
            except Exception as e: print(f"[CONFIRM] CRITICAL ERROR: {e}")
            finally: 
                if batch_id in ACTIVE_CONFIRMATIONS: ACTIVE_CONFIRMATIONS.remove(batch_id)
    thread = threading.Thread(target=task, args=(current_app._get_current_object().app_context(),)); thread.start()
    return jsonify({"status": "success", "message": "JOB STARTED - MONITORING"})

@main_bp.route('/api/download/csv/<batch_id>')
@login_required
def download_csv(batch_id):
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT filename, status FROM batches WHERE batch_id = ? AND user_id = ?", (batch_id, current_user.id)); batch_row = c.fetchone()
    if not batch_row: conn.close(); return jsonify({"error": "UNAUTHORIZED"}), 403
    if batch_row[1] == 'REFUNDED': conn.close(); return jsonify({"error": "ACCESS REVOKED: BATCH REFUNDED"}), 403
    download_name = f"{batch_id}.csv"
    if batch_row and batch_row[0] and '_' in batch_row[0]: download_name = batch_row[0].split('_', 1)[1]
    c.execute("SELECT id, from_name, to_name, tracking, created_at, address_to FROM history WHERE batch_id = ?", (batch_id,)); rows = c.fetchall(); conn.close()
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['No', 'Id', 'ClassService', 'FromName', 'ToName', 'TrackingId', 'TransDate', 'AddressTo'])
    for idx, row in enumerate(rows):
        try: fmt_date = (datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S") - timedelta(hours=5)).strftime("%m/%d/%Y %I:%M:%S %p")
        except: fmt_date = row[4]
        writer.writerow([idx + 1, row[0], "USPS Priority", row[1], row[2], row[3], fmt_date, row[5]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-disposition": f"attachment; filename={download_name}"})

@main_bp.route('/api/download/pdf/<batch_id>')
@login_required
def download_pdf(batch_id):
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT filename, status FROM batches WHERE batch_id = ? AND user_id = ?", (batch_id, current_user.id)); batch_row = c.fetchone()
    if not batch_row: return jsonify({"error": "UNAUTHORIZED ACCESS"}), 403
    if batch_row[1] == 'REFUNDED': return jsonify({"error": "ACCESS REVOKED: BATCH REFUNDED"}), 403
    clean_name = f"Batch_{batch_id}"
    if batch_row and batch_row[0] and '_' in batch_row[0]: clean_name = os.path.splitext(batch_row[0].split('_', 1)[1])[0]
    return send_from_directory(os.path.join(current_app.config['DATA_FOLDER'], 'pdfs'), f"{batch_id}.pdf", as_attachment=True, download_name=f"{clean_name}_{batch_id}.pdf")

@main_bp.route('/api/download/sample-csv')
@login_required
def download_sample_csv():
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(STRICT_HEADERS)
    writer.writerow(['1', 'Test LLC', '8823657928', '123 Jump St', 'Company LLC', '', 'New York', 'NY', '10001', 'Ben Dover', '9028439124', '123 Test St', '', '', 'New York', 'NY', '90001', '1', '1', '1', '1', 'desc', 'ref1', 'ref2', 'False', '1/12/2026'])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-disposition": "attachment; filename=Bulk_Template.csv"})

@main_bp.route('/api/addresses', methods=['GET'])
@login_required
def get_addresses_list():
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT * FROM sender_addresses WHERE user_id = ?", (current_user.id,))
    data = [{"id": r[0], "name": r[2], "company": r[3], "phone": r[4], "street1": r[5], "street2": r[6], "city": r[7], "state": r[8], "zip": r[9]} for r in c.fetchall()]
    conn.close(); return jsonify(data)

@main_bp.route('/api/addresses', methods=['POST'])
@login_required
def add_new_address():
    with address_lock: # Prevent race condition (Spam saving)
        d = request.json; conn = get_db_conn(); c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM sender_addresses WHERE user_id = ?", (current_user.id,)); 
        if c.fetchone()[0] >= 8: conn.close(); return jsonify({"error": "PROFILE LIMIT REACHED (8/8)"}), 400
        c.execute("INSERT INTO sender_addresses (user_id, name, company, street1, street2, city, state, zip, phone) VALUES (?,?,?,?,?,?,?,?,?)", 
                  (current_user.id, d['name'], d.get('company',''), d['street1'], d.get('street2', ''), d['city'], d['state'], d['zip'], d['phone']))
        conn.commit(); conn.close(); return jsonify({"status":"success"})

@main_bp.route('/api/addresses/<int:id>', methods=['DELETE'])
@login_required
def delete_single_address(id):
    conn = get_db_conn(); c = conn.cursor()
    c.execute("DELETE FROM sender_addresses WHERE id=? AND user_id=?", (id, current_user.id)); conn.commit(); conn.close()
    return jsonify({"status":"success"})

@main_bp.route('/api/addresses/all', methods=['DELETE'])
@login_required
def delete_all_user_addresses():
    conn = get_db_conn(); c = conn.cursor()
    c.execute("DELETE FROM sender_addresses WHERE user_id=?", (current_user.id,)); conn.commit(); conn.close()
    return jsonify({"status":"success"})

# --- SINGLE PURCHASE ROUTE ---
@main_bp.route('/api/purchase/single', methods=['POST'])
@login_required
@limiter.limit("20 per minute")
def purchase_single_label():
    data = request.json
    
    # 1. Validate Version Availability
    req_version = data.get('version', '95055')
    if not is_version_enabled(req_version):
        return jsonify({"error": f"Version {req_version} is currently disabled."}), 400

    # --- [ISSUE 1 FIX] READ TEMPLATE FROM REQUEST JSON ---
    # Previously, this might have been missing or read incorrectly.
    # We now explicitly look for 'template' in the data payload.
    req_template = data.get('template') 
    
    # Allow empty/null to default to Pitney, but if present, validate it.
    if req_template and req_template not in ['pitney_v2', 'stamps_v2', 'easypost_v2']:
         return jsonify({"error": "Invalid Template Selection"}), 400
    
    # Default fallback if user didn't send one (though frontend should)
    if not req_template: req_template = 'pitney_v2'

    # 2. Sender Logic (Saved vs Manual)
    sender_mode = data.get('sender_mode', 'saved')
    sender = {}
    
    if sender_mode == 'saved':
        s_id = data.get('sender_id')
        if not s_id: return jsonify({"error": "Please select a Sender Profile"}), 400
        saved = SenderAddress.get(s_id)
        if not saved or saved.user_id != current_user.id:
            return jsonify({"error": "Invalid Sender Profile"}), 400
        sender = {
            'FromName': sanitize_input(saved.name), 'CompanyFrom': sanitize_input(saved.company), 'PhoneFrom': sanitize_input(saved.phone),
            'Street1From': sanitize_input(saved.street1), 'Street2From': sanitize_input(saved.street2), 
            'CityFrom': sanitize_input(saved.city), 'StateFrom': sanitize_input(saved.state), 'PostalCodeFrom': sanitize_input(saved.zip)
        }
    else:
        # Manual Input
        sender = {
            'FromName': sanitize_input(data.get('s_name')), 'CompanyFrom': sanitize_input(data.get('s_company', '')), 
            'PhoneFrom': sanitize_input(data.get('s_phone')), 'Street1From': sanitize_input(data.get('s_street1')), 
            'Street2From': sanitize_input(data.get('s_street2', '')), 'CityFrom': sanitize_input(data.get('s_city')), 
            'StateFrom': sanitize_input(data.get('s_state')), 'PostalCodeFrom': sanitize_input(data.get('s_zip'))
        }
        # Basic Validation
        if not sender['FromName'] or not sender['Street1From'] or not sender['CityFrom'] or not sender['StateFrom'] or not sender['PostalCodeFrom']:
             return jsonify({"error": "Missing Required Sender Fields"}), 400

    # 3. Receiver Logic
    receiver = {
        'ToName': sanitize_input(data.get('r_name')), 'Company2': sanitize_input(data.get('r_company', '')), 
        'PhoneTo': sanitize_input(data.get('r_phone', '')), 'Street1To': sanitize_input(data.get('r_street1')), 
        'Street2To': sanitize_input(data.get('r_street2', '')), 'CityTo': sanitize_input(data.get('r_city')), 
        'StateTo': sanitize_input(data.get('r_state')), 'ZipTo': sanitize_input(data.get('r_zip'))
    }
    if not receiver['ToName'] or not receiver['Street1To'] or not receiver['CityTo'] or not receiver['StateTo'] or not receiver['ZipTo']:
        return jsonify({"error": "Missing Required Receiver Fields"}), 400

    # 4. Package Logic
    pkg = {
        'Weight': sanitize_input(data.get('weight', '1')),
        'Description': sanitize_input(data.get('description', '')),
        'Ref01': sanitize_input(data.get('ref1', '')),
        'Ref02': sanitize_input(data.get('ref2', '')),
        'Length': 10, 'Width': 6, 'Height': 4, # Defaults
        'Contains Hazard': 'False',
        'Shipment Date': datetime.now().strftime("%m/%d/%Y")
    }

    # 5. Pricing & Balance
    label_type = data.get('service', 'priority')
    price = get_price(current_user.id, label_type, req_version, current_user.price_per_label)
    
    # --- ATOMIC TRANSACTION ---
    with processing_lock:
        if not current_user.update_balance(-price):
            return jsonify({"error": "INSUFFICIENT FUNDS"}), 402
            
        try:
            # Generate Unique Batch ID
            batch_id = f"SINGLE_{random.randint(100000, 999999)}"
            
            # Create Data Row (Merging all dicts)
            row = {**sender, **receiver, **pkg, 'No': '1'}
            
            # Create CSV File for the Worker
            final_filename = f"{batch_id}.csv"
            save_path = os.path.join(current_app.config['DATA_FOLDER'], 'uploads', final_filename)
            
            # Use Pandas to write single row CSV efficiently
            df = pd.DataFrame([row])
            
            # Ensure all strict columns exist (Critical for Parser)
            for col in STRICT_HEADERS:
                if col not in df.columns: df[col] = ''
            
            df = df[STRICT_HEADERS] # Reorder to match template
            df.to_csv(save_path, index=False)
            
            # Insert into DB (Status: QUEUED)
            conn = get_db_conn(); c = conn.cursor()
            
            # --- IMPORTANT: Pass 'req_template' to DB ---
            c.execute("""
                INSERT INTO batches 
                (batch_id, user_id, filename, count, success_count, status, template, version, label_type, created_at, price) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (batch_id, current_user.id, final_filename, 1, 0, 'QUEUED', req_template, req_version, label_type, datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"), price))
            conn.commit(); conn.close()
            
            log_debug(f"Single Label Batch {batch_id} Queued")
            return jsonify({"status": "success", "batch_id": batch_id, "message": "Queued for Generation"})
            
        except Exception as e:
            current_user.update_balance(price) # Refund on crash
            log_debug(f"Single Purchase Error: {e}")
            return jsonify({"error": "System Error. Please try again."}), 500

# --- NEW: STATUS CHECKER FOR AUTO-DOWNLOAD ---
@main_bp.route('/api/batch/status/<batch_id>')
@login_required
def check_batch_status(batch_id):
    conn = get_db_conn(); c = conn.cursor()
    c.execute("SELECT status FROM batches WHERE batch_id = ? AND user_id = ?", (batch_id, current_user.id))
    row = c.fetchone()
    conn.close()
    
    if not row: return jsonify({"status": "NOT_FOUND"})
    
    # Map DB status to UI status
    status = row[0]
    # 'COMPLETED' is legacy, 'SUCCESS' implies history logic, but for batch table it's usually 'READY'
    # We check if it is done processing.
    if status in ['COMPLETED', 'SUCCESS', 'READY']: status = 'READY'
    
    return jsonify({"status": status})
